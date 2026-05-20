"""消息管理 API 路由"""
from __future__ import annotations

import json
import asyncio
import queue
import threading
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException, WebSocket, WebSocketDisconnect
from pydantic import BaseModel, Field

from config import load_config
from core.storage import build_blob_name, download_file_to_local, upload_file_to_storage
from openpyxl import Workbook, load_workbook
from db.auth_repository import resolve_user_from_authorization
from db.session_repository import (
    add_message,
    get_messages,
    get_session_by_id,
    add_session_file,
    get_session_files,
)
from core.agents.agent_d import run_agent_d_api
from service.agent_service import AgentService, get_selected_session_files_payload

router = APIRouter(prefix="/api/messages", tags=["消息管理"])


class SendMessageRequest(BaseModel):
    content: str = Field(..., description="消息内容")
    mode: Optional[str] = Field(
        default=None,
        description="工作模式；不传则使用会话在库中的 current_mode",
    )
    metadata: Optional[Dict[str, Any]] = Field(default=None, description="元数据")
    files: Optional[List[Dict[str, Any]]] = Field(
        default=None,
        description="随消息附带的数据文件（与 WebSocket 一致；不传则从会话已选文件读取）",
    )
    template_files: Optional[List[Dict[str, Any]]] = Field(
        default=None,
        description="随消息附带的模板文件",
    )


class MessageResponse(BaseModel):
    id: int
    session_id: str
    role: str
    content: str
    metadata: Optional[Dict[str, Any]] = None
    created_at: str


class SendMessageResponse(BaseModel):
    message_id: int
    content: str
    created_at: str


def _resolve_current_user(authorization: Optional[str], cfg):
    if not authorization:
        if cfg.auth.require_auth:
            raise HTTPException(status_code=401, detail="需要登录后访问")
        return None
    try:
        return resolve_user_from_authorization(authorization, cfg, required=True)
    except PermissionError as exc:
        raise HTTPException(status_code=401, detail=str(exc))


def _pick_table_filling_inputs(files: List[Dict[str, Any]], template_files: List[Dict[str, Any]]):
    """Pick one xlsx source and one template for direct table filling execution."""

    source_file = next((f for f in files if str(f.get("file_name", "")).lower().endswith(".xlsx")), None)
    docx_template = next((f for f in template_files if str(f.get("file_name", "")).lower().endswith(".docx")), None)
    xlsx_template = next((f for f in template_files if str(f.get("file_name", "")).lower().endswith(".xlsx")), None)
    template_file = docx_template or xlsx_template
    return source_file, template_file


def _resolve_file_reference(file_info: Dict[str, Any], cfg, session_id: str, kind: str) -> str:
    storage_key = str(file_info.get("storage_key") or "").strip()
    if not storage_key:
        return ""

    local_name = str(file_info.get("file_name") or storage_key).strip() or storage_key
    cache_path = Path(cfg.temp_dir) / "file_cache" / session_id / kind / local_name
    if cache_path.exists():
        return str(cache_path)

    if cfg.storage.enabled:
        try:
            return str(download_file_to_local(storage_key, cache_path, config=cfg))
        except Exception:
            pass

    # 本地临时文件路径检查：先尝试直接路径，再尝试相对 UPLOAD_DIR 的路径
    storage_path = Path(storage_key)
    if storage_path.is_absolute() and storage_path.exists():
        return str(storage_path)

    # 尝试相对于项目根目录的路径
    upload_dir = Path("workspace/uploads")
    relative_path = upload_dir / storage_key if not str(storage_key).startswith(str(upload_dir)) else storage_path
    if relative_path.exists():
        return str(relative_path)

    # 如果找不到，尝试作为相对路径直接使用（可能文件在其他位置）
    if storage_path.exists():
        return str(storage_path)

    # 返回原始路径，让调用方处理错误
    return storage_key


def _ensure_files_in_db(files: List[Dict[str, Any]], session_id: str, cfg, user_id: Optional[str]) -> List[Dict[str, Any]]:
    """
    确保文件记录在数据库中。
    如果文件是新的（storage_key 不在数据库中），则插入数据库并返回完整的文件信息（含 id）。
    """
    if not files:
        return []

    # 获取数据库中已有的文件
    try:
        db_files = get_session_files(session_id, config=cfg, user_id=user_id)
        db_keys = {getattr(f, "storage_key", None) or f.file_path: f for f in db_files}
    except Exception:
        db_files = []
        db_keys = {}

    result = []
    for f in files:
        storage_key = f.get("storage_key") or ""
        # 检查是否已在数据库中
        if storage_key in db_keys:
            db_file = db_keys[storage_key]
            result.append({
                "id": db_file.id,
                "file_id": db_file.id,
                "file_name": db_file.file_name,
                "storage_key": getattr(db_file, "storage_key", None) or db_file.file_path,
                "file_size": db_file.file_size,
                "file_type": db_file.file_type,
                "is_selected": True,
            })
        else:
            # 新文件，存入数据库
            file_name = f.get("file_name", "unnamed")
            file_type = f.get("file_type", "data")
            file_size = f.get("file_size", 0)

            session_file = add_session_file(
                session_id=session_id,
                file_name=file_name,
                file_type=file_type,
                file_path=storage_key,
                file_size=file_size,
                config=cfg,
                user_id=user_id,
                source="upload",
                role="source",
                storage_key=storage_key,
            )
            result.append({
                "id": session_file.id,
                "file_id": session_file.id,
                "file_name": session_file.file_name,
                "storage_key": getattr(session_file, "storage_key", None) or session_file.file_path,
                "file_size": session_file.file_size,
                "file_type": session_file.file_type,
                "is_selected": True,
            })

    return result


def _flatten_table_filling_response(response: Dict[str, Any]) -> Dict[str, Any]:
    """Convert run_agent_d_api result into the legacy frontend-friendly shape."""

    def _to_dict(obj: Any) -> Dict[str, Any]:
        if isinstance(obj, dict):
            return obj
        if isinstance(obj, str):
            try:
                parsed = json.loads(obj)
                return parsed if isinstance(parsed, dict) else {}
            except Exception:
                return {}
        if hasattr(obj, "model_dump"):
            try:
                dumped = obj.model_dump()
                return dumped if isinstance(dumped, dict) else {}
            except Exception:
                return {}
        if hasattr(obj, "dict"):
            try:
                dumped = obj.dict()
                return dumped if isinstance(dumped, dict) else {}
            except Exception:
                return {}
        if hasattr(obj, "__dict__"):
            try:
                dumped = dict(getattr(obj, "__dict__", {}))
                return dumped if isinstance(dumped, dict) else {}
            except Exception:
                return {}
        return {}

    data: Dict[str, Any] = {}
    resolved_input: Dict[str, Any] = {}
    resp_dict = _to_dict(response)
    if resp_dict:
        data = _to_dict(resp_dict.get("data"))
        resolved_input = _to_dict(resp_dict.get("resolved_input"))

        # 兼容旧形态：部分字段直接挂在 response 顶层
        if not data:
            top_level_candidates = (
                "status", "reason", "excel_path", "output_json", "template_output",
                "template_filled", "template_mapping", "multi_table_results",
                "total_rows", "matched_rows", "used_plan", "plan_source",
            )
            if any(k in resp_dict for k in top_level_candidates):
                data = resp_dict

    template_source = resolved_input.get("template") or data.get("template_source")

    output_json = data.get("output_json") or resolved_input.get("output_json")

    template_output = data.get("template_output") or data.get("output_template")
    if not template_output:
        template_output = resolved_input.get("output_template") or resolved_input.get("template_output")

    # 当 Agent 返回了 template_filled 但未显式给出 template_output 时，按默认命名规则回推路径
    if not template_output and bool(data.get("template_filled")):
        src_path = data.get("excel_path") or resolved_input.get("src")
        if src_path and template_source:
            try:
                src_obj = Path(str(src_path))
                tpl_suffix = Path(str(template_source)).suffix.lower() or ".xlsx"
                inferred = src_obj.parent / f"{src_obj.stem}_filled{tpl_suffix}"
                template_output = str(inferred)
            except Exception:
                template_output = None

    # output_json 兜底：按 AgentD 默认命名规则回推
    if not output_json:
        src_path = data.get("excel_path") or resolved_input.get("src")
        if src_path:
            try:
                src_obj = Path(str(src_path))
                output_json = str(src_obj.parent / f"{src_obj.stem}_filtered_rows.json")
            except Exception:
                output_json = None
    return {
        "success": bool(resp_dict.get("success", False)) if isinstance(resp_dict, dict) else False,
        "message": resp_dict.get("message", "") if isinstance(resp_dict, dict) else "",
        "status": data.get("status") if isinstance(data, dict) else None,
        "reason": data.get("reason") if isinstance(data, dict) else None,
        "multi_target_signal_level": data.get("multi_target_signal_level") if isinstance(data, dict) else None,
        "multi_target_mode_by_llm": data.get("multi_target_mode_by_llm") if isinstance(data, dict) else None,
        "excel_path": data.get("excel_path") if isinstance(data, dict) else None,
        "output_json": output_json,
        "total_rows": data.get("total_rows") if isinstance(data, dict) else 0,
        "matched_rows": data.get("matched_rows") if isinstance(data, dict) else 0,
        "used_plan": data.get("used_plan") if isinstance(data, dict) else None,
        "plan_source": data.get("plan_source") if isinstance(data, dict) else None,
        "template_filled": data.get("template_filled") if isinstance(data, dict) else False,
        "template_output": template_output,
        "template_source": template_source,
        "template_mapping": data.get("template_mapping") if isinstance(data, dict) else {},
        "multi_table_results": data.get("multi_table_results") if isinstance(data, dict) else None,
        "resolved_input": resolved_input,
    }


def _resolve_workspace_file(path_str: Optional[str]) -> Optional[Path]:
    """解析 Agent 输出的 workspace 相对路径或绝对路径，存在则返回 Path。"""
    if not path_str or not str(path_str).strip():
        return None
    p = Path(path_str)
    if not p.is_absolute():
        p = Path("workspace") / path_str
    return p if p.exists() else None


def _preview_rows_from_json_file(path: Path, max_rows: int) -> List[Dict[str, Any]]:
    try:
        raw = path.read_text(encoding="utf-8")
        data = json.loads(raw)
    except Exception as e:
        print(f"[WS] 读取 JSON 预览失败 {path}: {e}")
        return []
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)][:max_rows]
    if isinstance(data, dict):
        inner = data.get("rows") or data.get("filtered_rows") or data.get("entities")
        if isinstance(inner, list):
            return [x for x in inner if isinstance(x, dict)][:max_rows]
    return []


def _read_xlsx_preview_rows(xlsx_path: Path, max_rows: int = 40) -> List[Dict[str, Any]]:
    """从 xlsx 首行表头 + 若干数据行构建前端表格预览。"""
    rows: List[Dict[str, Any]] = []
    try:
        wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
        try:
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            first = next(it, None)
            if first is None:
                return rows
            headers: List[str] = []
            for j, c in enumerate(first):
                if c is not None and str(c).strip():
                    headers.append(str(c).strip())
                else:
                    headers.append(f"列{j + 1}")
            count = 0
            for row in it:
                count += 1
                if count > max_rows:
                    break
                d: Dict[str, Any] = {}
                for j, v in enumerate(row):
                    key = headers[j] if j < len(headers) else f"列{j + 1}"
                    d[key] = v
                rows.append(d)
        finally:
            wb.close()
    except Exception as e:
        print(f"[WS] _read_xlsx_preview_rows 失败 {xlsx_path}: {e}")
    return rows


def _build_table_filling_preview_rows(table_filling_data: Dict[str, Any], max_rows: int = 50) -> List[Dict[str, Any]]:
    """供 WebSocket / HTTP 在发送给前端前附加 previewData（混合模式表格子任务同路径）。"""
    oj = table_filling_data.get("output_json")
    if isinstance(oj, str) and oj.strip():
        jp = _resolve_workspace_file(oj)
        if jp and jp.suffix.lower() == ".json":
            got = _preview_rows_from_json_file(jp, max_rows)
            if got:
                return got
    tpl = table_filling_data.get("template_output")
    if isinstance(tpl, str) and tpl.strip():
        tp = _resolve_workspace_file(tpl)
        if tp and tp.suffix.lower() in (".xlsx", ".xlsm"):
            got = _read_xlsx_preview_rows(tp, max_rows=max_rows)
            if got:
                return got
    if isinstance(oj, str) and oj.strip():
        jp = _resolve_workspace_file(oj)
        if jp and jp.exists():
            got = _preview_rows_from_json_file(jp, max_rows)
            if got:
                return got
    return []


def _normalize_entity_extraction_response(raw_response: str) -> str:
    """将实体提取的原始 JSON 响应转换为简短摘要。"""
    if not isinstance(raw_response, str) or not raw_response.strip():
        return "实体提取完成，共提取 0 条数据"

    try:
        parsed = json.loads(raw_response)
    except Exception:
        return "实体提取完成"

    if not isinstance(parsed, dict):
        return "实体提取完成"

    entities = parsed.get("entities")
    count = len(entities) if isinstance(entities, list) else 0
    return f"实体提取完成，共提取 {count} 条数据"


def _message_to_dict(m) -> Dict[str, Any]:
    return {
        "id": m.id,
        "session_id": str(m.session_id),
        "role": m.role,
        "content": m.content,
        "metadata": m.metadata,
        "created_at": m.created_at.isoformat() + "Z" if m.created_at else "",
    }


def _persist_generated_files(session_id: str, cfg, user_id: Optional[str], payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    """将生成的文件持久化到 uploads 目录并记录到数据库，返回文件信息列表。"""
    uploads_dir = Path("workspace/uploads") / session_id
    uploads_dir.mkdir(parents=True, exist_ok=True)
    saved_files: List[Dict[str, Any]] = []
    candidate_paths = []
    for key in ("excel_path", "template_output", "output_json"):
        value = payload.get(key)
        if value:
            candidate_paths.append(str(value))
    for candidate in candidate_paths:
        try:
            path_obj = Path(candidate)
            if not path_obj.exists():
                continue
            dest_path = uploads_dir / path_obj.name
            import shutil
            shutil.copy2(path_obj, dest_path)
            storage_key = None
            try:
                storage_key = upload_file_to_storage(
                    dest_path,
                    config=cfg,
                    blob_name=build_blob_name(session_id, dest_path.name, prefix=cfg.storage.azure_blob_prefix),
                )
            except Exception:
                storage_key = None
            session_file = add_session_file(
                session_id=session_id,
                file_name=dest_path.name,
                file_type=dest_path.suffix.lower().lstrip(".") or "output",
                file_path=str(dest_path),
                file_size=dest_path.stat().st_size,
                config=cfg,
                user_id=user_id,
                source="generated",
                role="output",
                storage_key=storage_key,
            )
            saved_files.append({
                "file_id": session_file.id,
                "file_name": dest_path.name,
                "file_path": str(dest_path),
                "file_type": dest_path.suffix.lower().lstrip(".") or "output",
            })
        except Exception:
            continue
    return saved_files


def _save_entity_extraction_files(session_id: str, cfg, user_id: Optional[str], entities_json: str) -> List[Dict[str, Any]]:
    """将实体提取结果保存为 JSON 和 XLSX 文件，记录到数据库。"""
    uploads_dir = Path("workspace/uploads") / session_id
    uploads_dir.mkdir(parents=True, exist_ok=True)
    saved_files: List[Dict[str, Any]] = []

    try:
        parsed = json.loads(entities_json)
        entities = parsed.get("entities", []) if isinstance(parsed, dict) else []
    except Exception:
        entities = []

    # 1. 保存 JSON 文件
    json_name = f"extraction_result_{session_id[:8]}.json"
    json_path = uploads_dir / json_name
    try:
        json_path.write_text(entities_json, encoding="utf-8")
        storage_key_json = None
        try:
            storage_key_json = upload_file_to_storage(
                json_path,
                config=cfg,
                blob_name=build_blob_name(session_id, json_name, prefix=cfg.storage.azure_blob_prefix),
            )
        except Exception:
            storage_key_json = None

        sf_json = add_session_file(
            session_id=session_id,
            file_name=json_name,
            file_type="generated",
            file_path=str(json_path),
            file_size=json_path.stat().st_size,
            config=cfg,
            user_id=user_id,
            source="generated",
            role="output",
            storage_key=storage_key_json,
        )
        saved_files.append({"file_id": sf_json.id, "file_name": json_name, "file_path": str(json_path), "file_type": "json"})
        print(f"[WS] _save_entity_extraction_files: JSON保存成功 id={sf_json.id} type=json")
    except Exception as e:
        print(f"[WS] 保存 JSON 文件失败: {e}")

    # 2. 保存 XLSX 文件
    if entities:
        xlsx_name = f"extraction_result_{session_id[:8]}.xlsx"
        xlsx_path = uploads_dir / xlsx_name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "提取结果"

            # 表头
            if entities:
                headers = list(entities[0].keys())
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_idx, value=header)

                # 数据行
                for row_idx, entity in enumerate(entities, start=2):
                    for col_idx, header in enumerate(headers, start=1):
                        val = entity.get(header, "")
                        if isinstance(val, list) and len(val) >= 1:
                            val = val[0]
                        ws.cell(row=row_idx, column=col_idx, value=val)

            wb.save(str(xlsx_path))
            storage_key_xlsx = None
            try:
                storage_key_xlsx = upload_file_to_storage(
                    xlsx_path,
                    config=cfg,
                    blob_name=build_blob_name(session_id, xlsx_name, prefix=cfg.storage.azure_blob_prefix),
                )
            except Exception:
                storage_key_xlsx = None

            sf_xlsx = add_session_file(
                session_id=session_id,
                file_name=xlsx_name,
                file_type="generated",
                file_path=str(xlsx_path),
                file_size=xlsx_path.stat().st_size,
                config=cfg,
                user_id=user_id,
                source="generated",
                role="output",
                storage_key=storage_key_xlsx,
            )
            saved_files.append({"file_id": sf_xlsx.id, "file_name": xlsx_name, "file_path": str(xlsx_path), "file_type": "xlsx"})
            print(f"[WS] _save_entity_extraction_files: XLSX保存成功 id={sf_xlsx.id} type=xlsx")
        except Exception as e:
            print(f"[WS] 保存 XLSX 文件失败: {e}")

    print(f"[WS] _save_entity_extraction_files 返回: {saved_files}")
    return saved_files


def _save_table_filling_files(session_id: str, cfg, user_id: Optional[str], table_filling_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """将表格填表结果保存为 JSON 和 XLSX 文件，记录到数据库。"""
    uploads_dir = Path("workspace/uploads") / session_id
    uploads_dir.mkdir(parents=True, exist_ok=True)
    saved_files: List[Dict[str, Any]] = []

    # 1. 保存 JSON 文件（筛选结果）
    output_json = table_filling_data.get("output_json")
    if output_json:
        src_path = Path(output_json)
        if not src_path.is_absolute():
            src_path = Path("workspace") / output_json
        if src_path.exists():
            json_name = f"table_filling_result_{session_id[:8]}.json"
            dest_path = uploads_dir / json_name
            try:
                import shutil
                shutil.copy2(src_path, dest_path)
                storage_key = None
                try:
                    storage_key = upload_file_to_storage(
                        dest_path,
                        config=cfg,
                        blob_name=build_blob_name(session_id, json_name, prefix=cfg.storage.azure_blob_prefix),
                    )
                except Exception:
                    storage_key = None
                sf = add_session_file(
                    session_id=session_id,
                    file_name=json_name,
                    file_type="generated",
                    file_path=str(dest_path),
                    file_size=dest_path.stat().st_size,
                    config=cfg,
                    user_id=user_id,
                    source="generated",
                    role="output",
                    storage_key=storage_key,
                )
                saved_files.append({"file_id": sf.id, "file_name": json_name, "file_path": str(dest_path), "file_type": "json"})
                print(f"[WS] _save_table_filling_files: JSON保存成功 id={sf.id}")
            except Exception as e:
                print(f"[WS] _save_table_filling_files: 保存JSON失败: {e}")
        else:
            print(f"[WS] _save_table_filling_files: output_json不存在: {src_path}")

    # 2. 保存 XLSX 文件（填好的模板）
    template_output = table_filling_data.get("template_output")
    template_source = table_filling_data.get("template_source")
    expected_template_suffix = ""
    if template_source:
        try:
            expected_template_suffix = Path(str(template_source)).suffix.lower().lstrip(".")
        except Exception:
            expected_template_suffix = ""
    if template_output:
        src_path = Path(template_output)
        if not src_path.is_absolute():
            src_path = Path("workspace") / template_output
        if src_path.exists():
            ext = expected_template_suffix or src_path.suffix.lower().lstrip(".") or "xlsx"
            xlsx_name = f"table_filling_result_{session_id[:8]}.{ext}"
            dest_path = uploads_dir / xlsx_name
            try:
                import shutil
                shutil.copy2(src_path, dest_path)
                storage_key = None
                try:
                    storage_key = upload_file_to_storage(
                        dest_path,
                        config=cfg,
                        blob_name=build_blob_name(session_id, xlsx_name, prefix=cfg.storage.azure_blob_prefix),
                    )
                except Exception:
                    storage_key = None
                sf = add_session_file(
                    session_id=session_id,
                    file_name=xlsx_name,
                    file_type="generated",
                    file_path=str(dest_path),
                    file_size=dest_path.stat().st_size,
                    config=cfg,
                    user_id=user_id,
                    source="generated",
                    role="output",
                    storage_key=storage_key,
                )
                saved_files.append({"file_id": sf.id, "file_name": xlsx_name, "file_path": str(dest_path), "file_type": ext})
                print(f"[WS] _save_table_filling_files: XLSX保存成功 id={sf.id}")
            except Exception as e:
                print(f"[WS] _save_table_filling_files: 保存XLSX失败: {e}")
        else:
            print(f"[WS] _save_table_filling_files: template_output不存在: {src_path}")

    print(f"[WS] _save_table_filling_files 返回: {saved_files}")
    return saved_files


def _fallback_table_filling_generated_files(table_filling_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """当数据库保存失败或未命中时，使用 Agent 输出路径构造前端可下载文件列表。"""
    if not isinstance(table_filling_data, dict):
        return []
    out: List[Dict[str, Any]] = []
    template_source = table_filling_data.get("template_source")
    expected_template_suffix = ""
    if template_source:
        try:
            expected_template_suffix = Path(str(template_source)).suffix.lower().lstrip(".")
        except Exception:
            expected_template_suffix = ""
    output_json = table_filling_data.get("output_json")
    if output_json:
        p = str(output_json)
        out.append({
            "file_path": p,
            "file_name": Path(p).name or "table_filling_result.json",
            "file_type": "json",
        })
    template_output = table_filling_data.get("template_output")
    if template_output:
        p = str(template_output)
        ext = expected_template_suffix or Path(p).suffix.lower().lstrip(".") or "xlsx"
        out.append({
            "file_path": p,
            "file_name": Path(p).name or f"table_filling_result.{ext}",
            "file_type": ext,
        })
    return out


def _collect_new_generated_files(session_id: str, cfg, user_id: Optional[str], before_ids: set[int]) -> List[Dict[str, Any]]:
    generated: List[Dict[str, Any]] = []
    try:
        rows = get_session_files(session_id, config=cfg, user_id=user_id)
    except Exception as e:
        print(f"[WS] _collect_new_generated_files 查询失败: {e}")
        return generated
    print(f"[WS] _collect_new_generated_files: session_id={session_id}, before_ids={before_ids}, 总记录数={len(rows)}")
    for f in rows:
        print(f"[WS]   文件记录: id={f.id}, file_name={f.file_name}, role={getattr(f, 'role', '')}, file_type={getattr(f, 'file_type', '')}")
        if f.id in before_ids:
            print(f"[WS]   跳过(已在before_ids): {f.file_name}")
            continue
        if getattr(f, "role", "") != "output":
            print(f"[WS]   跳过(role非output): {f.file_name}")
            continue
        generated.append({
            "file_id": f.id,
            "file_name": f.file_name,
            "file_path": getattr(f, "file_path", ""),
            "file_type": getattr(f, "file_type", ""),
        })
    print(f"[WS] _collect_new_generated_files 返回: {generated}")
    return generated


@router.get("/{session_id}", response_model=List[MessageResponse])
async def get_messages_api(session_id: str, limit: int = 100, offset: int = 0, authorization: Optional[str] = Header(default=None)):
    """获取会话消息历史"""
    cfg = load_config()
    current_user = _resolve_current_user(authorization, cfg)
    messages = get_messages(session_id, limit=limit, offset=offset, config=cfg, user_id=current_user.id if current_user else None)
    return [_message_to_dict(m) for m in messages]


@router.post("/{session_id}", response_model=SendMessageResponse)
async def send_message(session_id: str, request: SendMessageRequest, authorization: Optional[str] = Header(default=None)):
    """
    发送消息并获取 AI 回复（非流式版本）
    用于简单场景或调试
    """
    cfg = load_config()
    current_user = _resolve_current_user(authorization, cfg)
    print(f"[API] POST /api/messages/{session_id} mode={request.mode or 'auto'} content={request.content[:80]}")
    
    # 检查会话是否存在
    session = get_session_by_id(session_id, config=cfg, user_id=current_user.id if current_user else None)
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")
    
    effective_mode = (request.mode or session.current_mode or "default_conversation").strip()

    user_meta: Dict[str, Any] = {**(request.metadata or {}), "mode": effective_mode}
    if request.files:
        user_meta["files"] = request.files
    if request.template_files:
        user_meta["template_files"] = request.template_files

    # 保存用户消息（含附件元数据，供前端展示「文件 + 文字」为一条消息）
    user_msg = add_message(
        session_id,
        "user",
        request.content,
        user_meta,
        config=cfg,
        user_id=current_user.id if current_user else None,
    )

    if request.files is not None or request.template_files is not None:
        db_data_files = list(request.files or [])
        db_template_files = list(request.template_files or [])
    else:
        db_data_files, db_template_files = get_selected_session_files_payload(session_id, cfg)

    # 确保临时文件在数据库中有记录
    db_data_files = _ensure_files_in_db(db_data_files, session_id, cfg, current_user.id if current_user else None)
    db_template_files = _ensure_files_in_db(db_template_files, session_id, cfg, current_user.id if current_user else None)

    # 表格填表走直达执行核，避免聊天/会话链路与 tests/test_d/run.py 的逻辑偏离。
    if effective_mode == "table_filling":
        source_file, template_file = _pick_table_filling_inputs(db_data_files, db_template_files)
        if source_file and template_file:
            request_table_targets = []
            if isinstance(request.metadata, dict):
                raw_targets = request.metadata.get("table_targets")
                if isinstance(raw_targets, list):
                    request_table_targets = raw_targets
            print(f"[API] 进入 table_filling 直达路径 session_id={session_id} source={source_file.get('file_name')} template={template_file.get('file_name')}")
            response = run_agent_d_api(
                src=_resolve_file_reference(source_file, cfg, session_id, "source"),
                prompt=request.content,
                template=_resolve_file_reference(template_file, cfg, session_id, "template"),
                output_json="",
                output_template="",
                allow_rule_fallback=True,
                table_targets=request_table_targets,
            )
            table_filling_data = _flatten_table_filling_response(response)
            print(f"[API] table_filling 响应 template_output={table_filling_data.get('template_output')} output_json={table_filling_data.get('output_json')} template_source={table_filling_data.get('template_source')}")
            saved = _save_table_filling_files(session_id, cfg, current_user.id if current_user else None, table_filling_data)
            print(f"[API] table_filling 保存文件: {saved}")
            if saved:
                table_filling_data["generated_files"] = saved
            preview_rows = _build_table_filling_preview_rows(table_filling_data)
            if preview_rows:
                table_filling_data["previewData"] = preview_rows
            ai_msg = add_message(
                session_id,
                "assistant",
                table_filling_data.get("message", ""),
                {"mode": effective_mode, "tableFillingData": table_filling_data},
                config=cfg,
                user_id=current_user.id if current_user else None,
            )
            return SendMessageResponse(
                message_id=ai_msg.id,
                content=ai_msg.content,
                created_at=ai_msg.created_at.isoformat() + "Z",
            )

    before_file_ids = {
        f.id for f in get_session_files(session_id, config=cfg, user_id=current_user.id if current_user else None)
    }

    # 调用 Agent 服务获取回复（必须与前端所选模式一致，否则恒走默认对话）
    agent_service = AgentService()
    response = await agent_service.chat(
        session_id,
        request.content,
        mode=effective_mode,
        files=db_data_files,
        template_files=db_template_files,
    )

    assistant_content = response
    assistant_meta: Dict[str, Any] = {"mode": effective_mode}
    if effective_mode == "entity_extraction":
        assistant_content = _normalize_entity_extraction_response(response)

    generated_files = _collect_new_generated_files(
        session_id,
        cfg,
        current_user.id if current_user else None,
        before_file_ids,
    )
    if generated_files:
        assistant_meta["generated_files"] = generated_files
    
    # 保存 AI 回复
    ai_msg = add_message(
        session_id,
        "assistant",
        assistant_content,
        assistant_meta,
        config=cfg,
        user_id=current_user.id if current_user else None,
    )
    
    return SendMessageResponse(
        message_id=ai_msg.id,
        content=ai_msg.content,
        created_at=ai_msg.created_at.isoformat() + "Z",
    )


# ============ WebSocket 流式聊天 ============

class ConnectionManager:
    """WebSocket 连接管理器"""
    
    def __init__(self):
        self.active_connections: Dict[str, WebSocket] = {}
    
    async def connect(self, websocket: WebSocket, session_id: str):
        await websocket.accept()
        self.active_connections[session_id] = websocket
    
    def disconnect(self, session_id: str):
        if session_id in self.active_connections:
            del self.active_connections[session_id]
    
    async def send_text(self, session_id: str, text: str):
        if session_id in self.active_connections:
            await self.active_connections[session_id].send_text(text)
    
    async def send_json(self, session_id: str, data: dict):
        if session_id not in self.active_connections:
            return
        try:
            await self.active_connections[session_id].send_json(data)
        except Exception:
            # WebSocket 已关闭，忽略发送失败
            pass


manager = ConnectionManager()


@router.websocket("/ws/{session_id}")
async def websocket_chat(websocket: WebSocket, session_id: str):
    """
    WebSocket 流式聊天（长连接模式）
    前端连接后可持续发送消息，后端保持连接循环处理。
    每次前端发送 JSON：{"content": "...", "mode": "...", "files": [...], "template_files": [...]}
    后端流式返回：{"type": "chunk", "content": "..."} 或 {"type": "done", "content": "..."}
    """
    cfg = load_config()
    current_user = None
    authorization = websocket.headers.get("authorization") or websocket.query_params.get("token")
    if authorization:
        try:
            current_user = resolve_user_from_authorization(authorization, cfg, required=True, allow_raw_token=True)
        except PermissionError as exc:
            await websocket.close(code=4401, reason=str(exc))
            return
    elif cfg.auth.require_auth:
        await websocket.close(code=4401, reason="需要登录后访问")
        return

    print(f"[API] WS /api/messages/ws/{session_id} connected")
    
    # 检查会话是否存在
    session = get_session_by_id(session_id, config=cfg, user_id=current_user.id if current_user else None)
    if not session:
        await websocket.close(code=4004, reason="会话不存在")
        return
    
    await manager.connect(websocket, session_id)
    print(f"[WS] 连接已建立 session_id={session_id}")
    
    try:
        # 保持连接循环，持续处理消息
        while True:
            try:
                # 等待接收消息（会阻塞在这里直到收到消息或连接关闭）
                data = await websocket.receive_json()
                print(f"[WS] 收到消息 session_id={session_id} mode={data.get('mode')} content={data.get('content','')[:50]}")
            except Exception:
                # 连接已关闭或出错，退出循环
                break
            
            user_content = data.get("content", "")
            # 显式 null/空串时回退到会话记录的模式，避免误走默认对话
            raw_mode = data.get("mode")
            mode = (raw_mode or session.current_mode or "default_conversation")
            if isinstance(mode, str):
                mode = mode.strip() or "default_conversation"
            
            # 优先使用前端传来的文件（和消息一起发送）
            # 这样用户可以随时切换勾选，文件跟随消息
            client_files = data.get("files") or []
            client_templates = data.get("template_files") or []
            client_table_targets = data.get("table_targets") or []

            # 实体提取模式
            if mode == "entity_extraction":
                db_data_files, db_template_files = get_selected_session_files_payload(session_id, cfg)
                db_path_map = {f.get('file_name'): f for f in db_data_files}
                db_tpl_path_map = {f.get('file_name'): f for f in db_template_files}
                if client_files:
                    files = []
                    for cf in client_files:
                        matched = db_path_map.get(cf.get('file_name'))
                        if matched:
                            files.append(matched)
                        elif cf.get('storage_key'):
                            files.append(cf)
                else:
                    files = db_data_files
                if client_templates:
                    template_files = []
                    for ct in client_templates:
                        matched = db_tpl_path_map.get(ct.get('file_name'))
                        if matched:
                            template_files.append(matched)
                        elif ct.get('storage_key'):
                            template_files.append(ct)
                else:
                    template_files = db_template_files
                # 确保临时文件在数据库中有记录
                files = _ensure_files_in_db(files, session_id, cfg, current_user.id if current_user else None)
                template_files = _ensure_files_in_db(template_files, session_id, cfg, current_user.id if current_user else None)
            elif mode == "table_filling":
                # 表格填表模式走直达执行核，保证与 tests/test_d/run.py 同构。
                db_data_files, db_template_files = get_selected_session_files_payload(session_id, cfg)
                db_path_map = {f.get('file_name'): f for f in db_data_files}
                db_tpl_path_map = {f.get('file_name'): f for f in db_template_files}
                if client_files:
                    files = []
                    for cf in client_files:
                        matched = db_path_map.get(cf.get('file_name'))
                        if matched:
                            files.append(matched)
                        elif cf.get('storage_key'):
                            files.append(cf)
                else:
                    files = db_data_files
                if client_templates:
                    template_files = []
                    for ct in client_templates:
                        matched = db_tpl_path_map.get(ct.get('file_name'))
                        if matched:
                            template_files.append(matched)
                        elif ct.get('storage_key'):
                            template_files.append(ct)
                else:
                    template_files = db_template_files
                # 确保临时文件在数据库中有记录
                files = _ensure_files_in_db(files, session_id, cfg, current_user.id if current_user else None)
                template_files = _ensure_files_in_db(template_files, session_id, cfg, current_user.id if current_user else None)

                source_file, template_file = _pick_table_filling_inputs(files, template_files)
                if source_file and template_file:
                    print(f"[API] WS table_filling 开始处理 session_id={session_id} source={source_file.get('file_name')} template={template_file.get('file_name')}")
                    user_meta: Dict[str, Any] = {"mode": mode}
                    if files:
                        user_meta["files"] = files
                    if template_files:
                        user_meta["template_files"] = template_files
                    add_message(session_id, "user", user_content, user_meta, config=cfg, user_id=current_user.id if current_user else None)
                    print(f"[WS] 发送 table_filling start type=start mode={mode}")
                    await manager.send_json(session_id, {"type": "start", "mode": mode})
                    print(f"[WS] 调用 run_agent_d_api...")
                    response = await asyncio.to_thread(
                        run_agent_d_api,
                        src=_resolve_file_reference(source_file, cfg, session_id, "source"),
                        prompt=user_content,
                        template=_resolve_file_reference(template_file, cfg, session_id, "template"),
                        output_json="",
                        output_template="",
                        allow_rule_fallback=True,
                        table_targets=client_table_targets if isinstance(client_table_targets, list) else None,
                    )
                    print(f"[API] WS run_agent_d_api 完成, 响应长度={len(str(response))}")
                    if isinstance(response, dict):
                        data_obj = response.get("data") if isinstance(response.get("data"), dict) else {}
                        resolved_obj = response.get("resolved_input") if isinstance(response.get("resolved_input"), dict) else {}
                        print(
                            f"[API] WS run_agent_d_api 摘要: success={response.get('success')} "
                            f"status={data_obj.get('status')} reason={data_obj.get('reason')}"
                        )
                        print(
                            f"[API] WS run_agent_d_api 原始字段: "
                            f"data.output_json={data_obj.get('output_json')} "
                            f"data.template_output={data_obj.get('template_output')} "
                            f"resolved.output_json={resolved_obj.get('output_json')} "
                            f"resolved.output_template={resolved_obj.get('output_template')}"
                        )
                    table_filling_data = _flatten_table_filling_response(response)
                    if not table_filling_data.get("success", False):
                        error_msg = table_filling_data.get("message") or "表格填表失败"
                        print(
                            f"[API] WS table_filling 失败: message={error_msg}; "
                            f"reason={table_filling_data.get('reason')}; "
                            f"signal={table_filling_data.get('multi_target_signal_level')}; "
                            f"llm_mode={table_filling_data.get('multi_target_mode_by_llm')}"
                        )
                        await manager.send_json(
                            session_id,
                            {
                                "type": "error",
                                "mode": mode,
                                "result_type": "table_filling",
                                "message": error_msg,
                                "data": table_filling_data,
                            },
                        )
                        add_message(
                            session_id,
                            "assistant",
                            error_msg,
                            {"mode": mode, "tableFillingData": table_filling_data},
                            config=cfg,
                            user_id=current_user.id if current_user else None,
                        )
                        continue
                    print(f"[API] WS table_filling_data keys={list(table_filling_data.keys())}")
                    print(f"[API] WS table_filling_data template_output={table_filling_data.get('template_output')} output_json={table_filling_data.get('output_json')} template_source={table_filling_data.get('template_source')}")
                    saved = _save_table_filling_files(session_id, cfg, current_user.id if current_user else None, table_filling_data)
                    print(f"[API] WS _save_table_filling_files 返回: {saved}")
                    if saved:
                        table_filling_data["generated_files"] = saved
                    preview_rows = _build_table_filling_preview_rows(table_filling_data)
                    if preview_rows:
                        table_filling_data["previewData"] = preview_rows
                        print(f"[API] WS table_filling 附加 previewData 行数={len(preview_rows)}")
                    full_response = json.dumps(table_filling_data, ensure_ascii=False)
                    print(f"[API] WS 发送 table_filling chunk, 内容长度={len(full_response)}")
                    await manager.send_json(session_id, {"type": "chunk", "content": full_response, "result_type": "table_filling"})
                    add_message(session_id, "assistant", table_filling_data.get("message", ""), {"mode": mode, "tableFillingData": table_filling_data}, config=cfg, user_id=current_user.id if current_user else None)
                    print(f"[API] WS 发送 table_filling done")
                    done_payload: Dict[str, Any] = {
                        "type": "done",
                        "mode": mode,
                        "result_type": "table_filling",
                        "table_filling_data": table_filling_data,
                    }
                    final_files = saved if saved else _fallback_table_filling_generated_files(table_filling_data)
                    if final_files:
                        done_payload["generated_files"] = final_files
                    await manager.send_json(session_id, done_payload)
                    continue
            elif client_files or client_templates:
                files = client_files
                template_files = client_templates
            else:
                files, template_files = get_selected_session_files_payload(session_id, cfg)

            # 确保临时文件在数据库中有记录
            files = _ensure_files_in_db(files, session_id, cfg, current_user.id if current_user else None)
            template_files = _ensure_files_in_db(template_files, session_id, cfg, current_user.id if current_user else None)

            user_meta: Dict[str, Any] = {"mode": mode}
            if files:
                user_meta["files"] = files
            if template_files:
                user_meta["template_files"] = template_files

            # 保存用户消息（含附件元数据）
            add_message(session_id, "user", user_content, user_meta, config=cfg, user_id=current_user.id if current_user else None)
            
            # 发送开始信号
            print(f"[WS] 发送 start type=start mode={mode} session_id={session_id}")
            await manager.send_json(session_id, {"type": "start", "mode": mode})

            before_file_ids = {
                f.id for f in get_session_files(session_id, config=cfg, user_id=current_user.id if current_user else None)
            }

            # 进度队列：线程安全信令，规避 run_coroutine_threadsafe 在主 loop 阻塞时无法执行的问题
            progress_queue: queue.Queue = queue.Queue()

            def progress_callback(completed: int, total: int, message: str):
                percent = int(completed / total * 100) if total > 0 else 0
                progress_queue.put_nowait({
                    "type": "progress",
                    "progress": percent,
                    "message": message,
                })

            # 在后台线程跑提取，主 loop 保持空闲以处理 WebSocket
            agent_service = AgentService()
            full_response = ""

            async def drain_queue():
                """每 50ms 把队列中的进度消息发往 WebSocket"""
                while not progress_queue.empty():
                    try:
                        msg = progress_queue.get_nowait()
                        # _done 标记仅用于退出循环，不发往 WebSocket
                        if msg.get("_done"):
                            continue
                        await manager.send_json(session_id, msg)
                    except queue.Empty:
                        break

            async def extraction_task():
                nonlocal full_response
                try:
                    async for chunk in agent_service.chat_stream(
                        session_id,
                        user_content,
                        mode,
                        files=files,
                        template_files=template_files,
                        progress_callback=progress_callback if mode in ("entity_extraction", "table_filling") else None,
                    ):
                        full_response += chunk
                        if mode == "entity_extraction":
                            print(f"[WS] 发送 entity_extraction chunk, 内容长度={len(chunk)} session_id={session_id}")
                            await manager.send_json(session_id, {"type": "chunk", "content": chunk, "result_type": "entity_extraction"})
                        elif mode == "table_filling":
                            print(f"[WS] 发送 table_filling chunk, 内容长度={len(chunk)} session_id={session_id}")
                            await manager.send_json(session_id, {"type": "chunk", "content": chunk, "result_type": "table_filling"})
                        else:
                            print(f"[WS] 发送普通 chunk, 内容长度={len(chunk)} session_id={session_id}")
                            await manager.send_json(session_id, {"type": "chunk", "content": chunk})
                finally:
                    progress_queue.put_nowait({"_done": True})

            # 并发：提取跑后台 + 主 loop 轮询进度队列
            ext_task = asyncio.create_task(extraction_task())
            while not ext_task.done():
                await drain_queue()
                await asyncio.sleep(0.05)
            await drain_queue()

            assistant_content = full_response
            assistant_meta: Dict[str, Any] = {"mode": mode}
            extraction_files: List[Dict[str, Any]] = []
            if mode == "entity_extraction":
                assistant_content = _normalize_entity_extraction_response(full_response)
                extraction_files = _save_entity_extraction_files(
                    session_id,
                    cfg,
                    current_user.id if current_user else None,
                    full_response,
                )

            generated_files = _collect_new_generated_files(
                session_id,
                cfg,
                current_user.id if current_user else None,
                before_file_ids,
            )
            if extraction_files:
                assistant_meta["generated_files"] = extraction_files
            elif generated_files:
                assistant_meta["generated_files"] = generated_files
            elif mode == "table_filling":
                try:
                    parsed = json.loads(full_response)
                except Exception:
                    parsed = None
                if isinstance(parsed, dict):
                    fallback_files = _fallback_table_filling_generated_files(parsed)
                    if fallback_files:
                        assistant_meta["generated_files"] = fallback_files
                    assistant_meta["tableFillingData"] = parsed

            add_message(
                session_id,
                "assistant",
                assistant_content,
                assistant_meta,
                config=cfg,
                user_id=current_user.id if current_user else None,
            )
            # done 消息带上 generated_files，前端据此显示下载按钮
            done_payload: Dict[str, Any] = {"type": "done", "mode": mode}
            final_files = extraction_files if extraction_files else generated_files
            if mode == "table_filling" and not final_files:
                tf_data = assistant_meta.get("tableFillingData")
                if isinstance(tf_data, dict):
                    final_files = _fallback_table_filling_generated_files(tf_data)
                    done_payload["result_type"] = "table_filling"
                    done_payload["table_filling_data"] = tf_data
            if final_files:
                done_payload["generated_files"] = final_files
            print(f"[WS] done_payload generated_files: {done_payload.get('generated_files')}")
            await manager.send_json(session_id, done_payload)
            print(f"[WS] 发送 done session_id={session_id} extraction_files={extraction_files} generated_files={generated_files}")
            
    except WebSocketDisconnect:
        print(f"[WS] WebSocketDisconnect 正常断开 session_id={session_id}")
        pass  # 正常断开
    except Exception as e:
        print(f"[WS] Exception: {e} session_id={session_id}")
        import traceback; traceback.print_exc()
        await manager.send_json(session_id, {"type": "error", "message": str(e)})
    finally:
        manager.disconnect(session_id)
