"""XLSX adapter for executing spreadsheet actions with openpyxl."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from utils.document_reader import ExcelReader

from .standard_style import get_standard_style_preset


@dataclass
class ActionExecutionResult:
    action_type: str
    success: bool
    message: str
    details: Dict[str, Any]


class XlsxAdapter:
    """Adapter that applies action items to xlsx files."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = load_workbook(file_path)
        self.execution_log: List[ActionExecutionResult] = []

    def apply_action(self, action: Dict[str, Any]) -> ActionExecutionResult:
        action_type = action.get("action_type", "")
        params = action.get("params", {}) or {}
        target = action.get("target", {}) or {}

        handler_map = {
            "batch_format": self._apply_batch_format,
            "unify_style": self._apply_unify_style,
            "extract_content": self._apply_extract_content,
            "reorder_paragraphs": self._apply_reorder_rows,
            "auto_column_width": self._apply_auto_column_width,
            "freeze_header_row": self._apply_freeze_header_row,
            "replace_text": self._apply_replace_text,
        }

        handler = handler_map.get(action_type)
        if handler is None:
            result = ActionExecutionResult(action_type, False, f"XLSX 适配器暂不支持动作: {action_type}", {})
            self.execution_log.append(result)
            return result

        try:
            details = handler(target, params)
            result = ActionExecutionResult(action_type, True, "执行成功", details)
        except Exception as e:
            result = ActionExecutionResult(action_type, False, f"执行失败: {e}", {})

        self.execution_log.append(result)
        return result

    def save(self, output_path: str) -> str:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(path)
        return str(path)

    def _resolve_sheet(self, target: Dict[str, Any], params: Dict[str, Any]) -> Worksheet:
        sheet_name = target.get("sheet") or params.get("sheet")
        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"工作表不存在: {sheet_name}")
            return self.workbook[sheet_name]
        return self.workbook[self.workbook.sheetnames[0]]

    def _resolve_range(self, ws: Worksheet, target: Dict[str, Any], params: Dict[str, Any]) -> str:
        range_ref = target.get("range") or params.get("range")
        if isinstance(range_ref, str) and ":" in range_ref:
            return range_ref

        max_row = max(ws.max_row, 1)
        max_col = max(ws.max_column, 1)
        return f"A1:{get_column_letter(max_col)}{max_row}"

    def _iter_cells(self, ws: Worksheet, range_ref: str) -> Iterable[Cell]:
        for row in ws[range_ref]:
            for cell in row:
                yield cell

    @staticmethod
    def _normalize_color(color: Any) -> Optional[str]:
        if color is None:
            return None
        s = str(color).strip().upper().replace("#", "")
        if not s:
            return None
        if len(s) == 6:
            return f"FF{s}"
        if len(s) == 8:
            return s
        return None

    def _build_font(self, base: Font, params: Dict[str, Any]) -> Font:
        color = self._normalize_color(params.get("font_color"))
        return Font(
            name=params.get("font_name", base.name),
            size=float(params.get("font_size", base.size or 11)),
            bold=bool(params.get("bold", base.bold)),
            italic=bool(params.get("italic", base.italic)),
            underline=params.get("underline", base.underline),
            color=color or base.color,
        )

    def _build_alignment(self, base: Alignment, params: Dict[str, Any]) -> Alignment:
        return Alignment(
            horizontal=params.get("horizontal", base.horizontal),
            vertical=params.get("vertical", base.vertical),
            wrap_text=bool(params.get("wrap_text", base.wrap_text)),
            text_rotation=int(params.get("text_rotation", base.text_rotation or 0)),
            shrink_to_fit=bool(params.get("shrink_to_fit", base.shrink_to_fit)),
            indent=int(params.get("indent", base.indent or 0)),
        )

    def _build_border(self, params: Dict[str, Any]) -> Border:
        style = str(params.get("border_style", "thin"))
        color = self._normalize_color(params.get("border_color")) or "FF000000"
        side = Side(style=style, color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    def _parse_column_spec(value: Any) -> List[str]:
        if value is None:
            return []
        if isinstance(value, str):
            parts = re.split(r"[，,\s]+", value.strip())
            return [p.upper() for p in parts if p]
        if isinstance(value, list):
            return [str(v).strip().upper() for v in value if str(v).strip()]
        return []

    def _apply_column_width(self, ws: Worksheet, params: Dict[str, Any]) -> Dict[str, float]:
        updated: Dict[str, float] = {}
        width_cfg = params.get("column_width")

        if isinstance(width_cfg, (int, float)):
            columns = self._parse_column_spec(params.get("columns"))
            for col in columns:
                ws.column_dimensions[col].width = float(width_cfg)
                updated[col] = float(width_cfg)
            return updated

        if isinstance(width_cfg, dict):
            for col, width in width_cfg.items():
                col_letter = str(col).strip().upper()
                ws.column_dimensions[col_letter].width = float(width)
                updated[col_letter] = float(width)
            return updated

        return updated

    def _apply_batch_format(self, target: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
        ws = self._resolve_sheet(target, params)
        range_ref = self._resolve_range(ws, target, params)

        applied = 0
        border = self._build_border(params) if params.get("apply_border", True) else None

        for cell in self._iter_cells(ws, range_ref):
            if cell.value is None and not params.get("format_empty", False):
                continue
            cell.font = self._build_font(cell.font or Font(), params)
            cell.alignment = self._build_alignment(cell.alignment or Alignment(), params)
            if border is not None:
                cell.border = border
            applied += 1

        updated_cols = self._apply_column_width(ws, params)
        return {
            "sheet": ws.title,
            "range": range_ref,
            "formatted_cells": applied,
            "updated_column_widths": updated_cols,
        }

    def _apply_unify_style(self, target: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
        ws = self._resolve_sheet(target, params)
        max_row = max(ws.max_row, 1)
        max_col = max(ws.max_column, 1)
        preset = get_standard_style_preset("xlsx") if str(params.get("style_preset", "")).lower() == "standard" else {}
        merged = {**preset, **params}

        border = self._build_border({"border_style": merged.get("border_style", "thin"), "border_color": merged.get("border_color", "000000")})

        header_font = Font(name=str(merged.get("header_font_name", "Calibri")), size=float(merged.get("header_font_size", 11)), bold=True)
        body_font = Font(name=str(merged.get("body_font_name", "Calibri")), size=float(merged.get("body_font_size", 11)), bold=False)
        header_align = Alignment(horizontal=str(merged.get("header_horizontal", "center")), vertical="center", wrap_text=True)
        body_align = Alignment(horizontal=str(merged.get("body_horizontal", "left")), vertical="center", wrap_text=True)

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if row_idx == 1:
                    cell.font = header_font
                    cell.alignment = header_align
                else:
                    cell.font = body_font
                    cell.alignment = body_align
                cell.border = border

        return {
            "sheet": ws.title,
            "rows": max_row,
            "cols": max_col,
            "strategy": merged.get("strategy", "standard"),
            "style_preset": merged.get("style_preset", "standard"),
        }

    def _header_map(self, ws: Worksheet, header_row: int = 1) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col).value
            if cell is None:
                continue
            key = str(cell).strip()
            if key:
                mapping[key] = col
        return mapping

    def _resolve_col_index(self, ws: Worksheet, col_spec: Any, header_row: int = 1) -> Optional[int]:
        if col_spec is None:
            return None
        if isinstance(col_spec, int):
            return col_spec

        text = str(col_spec).strip()
        if not text:
            return None

        if text.isdigit():
            return int(text)

        if re.match(r"^[A-Za-z]+$", text):
            return column_index_from_string(text.upper())

        return self._header_map(ws, header_row).get(text)

    @staticmethod
    def _evaluate_condition(cell_value: Any, op: str, expected: Any) -> bool:
        op = (op or "==").strip()
        if op in ("contains", "in"):
            return str(expected) in str(cell_value)

        try:
            left_num = float(cell_value)
            right_num = float(expected)
            if op == "==":
                return left_num == right_num
            if op == "!=":
                return left_num != right_num
            if op == ">":
                return left_num > right_num
            if op == ">=":
                return left_num >= right_num
            if op == "<":
                return left_num < right_num
            if op == "<=":
                return left_num <= right_num
        except Exception:
            left = "" if cell_value is None else str(cell_value)
            right = "" if expected is None else str(expected)
            if op == "==":
                return left == right
            if op == "!=":
                return left != right
            if op == "contains":
                return right in left

        return False

    def _normalize_where(self, where: Any) -> Optional[Dict[str, Any]]:
        if isinstance(where, dict):
            if "column" in where and "value" in where:
                return {
                    "column": where.get("column"),
                    "op": where.get("op", "=="),
                    "value": where.get("value"),
                }
            return None

        if isinstance(where, str):
            m = re.match(r"^\s*(.+?)\s*(==|!=|>=|<=|>|<)\s*(.+?)\s*$", where)
            if m:
                return {"column": m.group(1).strip(), "op": m.group(2), "value": m.group(3).strip()}
            m2 = re.match(r"^\s*(.+?)\s*包含\s*(.+?)\s*$", where)
            if m2:
                return {"column": m2.group(1).strip(), "op": "contains", "value": m2.group(2).strip()}
        return None

    def _apply_extract_content(self, target: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
        ws = self._resolve_sheet(target, params)
        header_row = int(params.get("header_row", 1))
        where = self._normalize_where(params.get("where") or params.get("condition"))

        columns_spec = params.get("columns")
        if columns_spec is None:
            columns_spec = [ws.cell(row=header_row, column=i).value for i in range(1, ws.max_column + 1)]

        selected_cols: List[Tuple[str, int]] = []
        for spec in columns_spec:
            idx = self._resolve_col_index(ws, spec, header_row=header_row)
            if idx is None:
                continue
            name = ws.cell(row=header_row, column=idx).value
            selected_cols.append((str(name or get_column_letter(idx)), idx))

        where_col_idx = None
        where_op = "=="
        where_value = None
        if where:
            where_col_idx = self._resolve_col_index(ws, where.get("column"), header_row=header_row)
            where_op = str(where.get("op", "=="))
            where_value = where.get("value")

        rows: List[Dict[str, Any]] = []
        stat_rows: List[List[str]] = []

        for r in range(header_row + 1, ws.max_row + 1):
            if where_col_idx is not None:
                cell_val = ws.cell(row=r, column=where_col_idx).value
                if not self._evaluate_condition(cell_val, where_op, where_value):
                    continue

            row_item: Dict[str, Any] = {}
            stat_row: List[str] = []
            for name, idx in selected_cols:
                val = ws.cell(row=r, column=idx).value
                row_item[name] = val
                stat_row.append("" if val is None else str(val))
            if row_item:
                rows.append(row_item)
                stat_rows.append(stat_row)

        headers = [name for name, _ in selected_cols]
        stats_text = ""
        if headers and stat_rows:
            reader = ExcelReader()
            stats_text = reader._compute_statistics(headers, stat_rows)

        return {
            "sheet": ws.title,
            "columns": headers,
            "where": where,
            "rows": rows,
            "row_count": len(rows),
            "stats": stats_text,
        }

    def _apply_reorder_rows(self, target: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
        ws = self._resolve_sheet(target, params)
        header_row = int(params.get("header_row", 1))

        sort_by = params.get("sort_by")
        if sort_by:
            sort_col = self._resolve_col_index(ws, sort_by, header_row=header_row)
            if sort_col is None:
                return {"moved": False, "reason": f"无效排序列: {sort_by}"}

            start_row = int(params.get("start_row", header_row + 1))
            end_row = int(params.get("end_row", ws.max_row))
            descending = str(params.get("order", "asc")).lower() == "desc"

            rows_data: List[Tuple[int, List[Any]]] = []
            for r in range(start_row, end_row + 1):
                row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                rows_data.append((r, row_values))

            def _key(item: Tuple[int, List[Any]]) -> Tuple[int, Any]:
                v = item[1][sort_col - 1]
                if v is None:
                    return (1, "")
                try:
                    return (0, float(v))
                except Exception:
                    return (0, str(v))

            sorted_rows = sorted(rows_data, key=_key, reverse=descending)
            for idx, (_, values) in enumerate(sorted_rows):
                row_no = start_row + idx
                for c, val in enumerate(values, start=1):
                    ws.cell(row=row_no, column=c).value = val

            return {
                "moved": True,
                "mode": "sort",
                "sheet": ws.title,
                "sort_by": sort_by,
                "order": "desc" if descending else "asc",
                "affected_rows": len(rows_data),
            }

        move_range = params.get("move_range")
        to_row = params.get("to_row")
        if isinstance(move_range, str) and ":" in move_range and to_row is not None:
            start_s, end_s = move_range.split(":", 1)
            start_row = int(start_s)
            end_row = int(end_s)
            to_row_idx = int(to_row)

            block = [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)] for r in range(start_row, end_row + 1)]
            others = [
                [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)
                if not (start_row <= r <= end_row)
            ]

            insert_at = max(0, min(len(others), to_row_idx - 1))
            reordered = others[:insert_at] + block + others[insert_at:]

            for r in range(1, len(reordered) + 1):
                for c, val in enumerate(reordered[r - 1], start=1):
                    ws.cell(row=r, column=c).value = val

            return {
                "moved": True,
                "mode": "move_region",
                "sheet": ws.title,
                "from": move_range,
                "to_row": to_row_idx,
                "affected_rows": len(block),
            }

        from_idx = params.get("from")
        to_idx = params.get("to")
        if from_idx is None or to_idx is None:
            return {"moved": False, "reason": "缺少排序或重排参数"}

        from_row = int(from_idx) + header_row
        to_row = int(to_idx) + header_row
        if from_row < header_row + 1 or to_row < header_row + 1 or from_row > ws.max_row or to_row > ws.max_row:
            return {"moved": False, "reason": "索引越界", "from": from_idx, "to": to_idx}

        row_values = [ws.cell(row=from_row, column=c).value for c in range(1, ws.max_column + 1)]
        if from_row < to_row:
            for r in range(from_row, to_row):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).value = ws.cell(row=r + 1, column=c).value
        elif from_row > to_row:
            for r in range(from_row, to_row, -1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).value = ws.cell(row=r - 1, column=c).value

        for c, val in enumerate(row_values, start=1):
            ws.cell(row=to_row, column=c).value = val

        return {"moved": True, "mode": "move_row", "from": from_idx, "to": to_idx, "sheet": ws.title}

    def _apply_auto_column_width(self, target: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
        ws = self._resolve_sheet(target, params)
        padding = int(params.get("padding", 2))
        min_width = float(params.get("min_width", 8))
        max_width = float(params.get("max_width", 60))

        updated: Dict[str, float] = {}
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            width = min(max_width, max(min_width, float(max_len + padding)))
            ws.column_dimensions[col_letter].width = width
            updated[col_letter] = width

        return {"sheet": ws.title, "updated_column_widths": updated}

    def _apply_freeze_header_row(self, target: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
        ws = self._resolve_sheet(target, params)
        header_row = int(params.get("header_row", 1))
        ws.freeze_panes = f"A{header_row + 1}"
        return {"sheet": ws.title, "freeze_panes": ws.freeze_panes}

    def _apply_replace_text(self, target: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
        ws = self._resolve_sheet(target, params)
        find_text = str(params.get("find", ""))
        replace_text = str(params.get("replace", ""))
        if not find_text:
            return {"find": find_text, "replace": replace_text, "replaced": 0}

        count = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and find_text in cell.value:
                    cell.value = cell.value.replace(find_text, replace_text)
                    count += 1

        return {"sheet": ws.title, "find": find_text, "replace": replace_text, "replaced": count}
