"""XLSX 适配器独立动作测试（3个样本表）。"""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.agents.agent_a.xlsx_adapter import XlsxAdapter


def _create_sample_sales(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["姓名", "地区", "销售额", "状态"])
    ws.append(["张三", "华东", 120, "在售"])
    ws.append(["李四", "华北", 80, "停售"])
    ws.append(["王五", "华南", 150, "在售"])
    wb.save(path)
    return path


def _create_sample_inventory(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["SKU", "品类", "库存", "单价"])
    ws.append(["A-01", "药品", 20, 13.5])
    ws.append(["B-02", "器械", 8, 120.0])
    ws.append(["C-03", "耗材", 50, 6.8])
    wb.save(path)
    return path


def _create_sample_region(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Region"
    ws.append(["区域", "月度", "完成率"])
    ws.append(["华东", "1月", 0.92])
    ws.append(["华北", "1月", 0.88])
    ws.append(["华南", "1月", 0.95])
    ws.append(["华中", "1月", 0.9])
    wb.save(path)
    return path


def test_xlsx_case_01_batch_format_font_alignment_border_column_width(tmp_path: Path):
    p = _create_sample_sales(tmp_path / "sales.xlsx")
    a = XlsxAdapter(str(p))
    result = a.apply_action(
        {
            "action_type": "batch_format",
            "target": {"sheet": "Sales", "range": "A1:D4"},
            "params": {
                "font_name": "Calibri",
                "font_size": 12,
                "bold": True,
                "horizontal": "center",
                "vertical": "center",
                "apply_border": True,
                "column_width": {"A": 14, "B": 12},
            },
        }
    )
    out = tmp_path / "sales_fmt.xlsx"
    a.save(str(out))

    wb = load_workbook(out)
    ws = wb["Sales"]
    assert result.success is True
    assert ws["A2"].font.bold is True
    assert ws["A2"].font.size == 12
    assert ws["A2"].alignment.horizontal == "center"
    assert ws["A2"].border.left.style == "thin"
    assert ws.column_dimensions["A"].width == 14
    assert ws.column_dimensions["B"].width == 12


def test_xlsx_case_02_batch_format_idempotent(tmp_path: Path):
    p = _create_sample_sales(tmp_path / "sales.xlsx")
    a = XlsxAdapter(str(p))
    action = {
        "action_type": "batch_format",
        "target": {"sheet": "Sales", "range": "A1:D4"},
        "params": {
            "font_name": "Calibri",
            "font_size": 11,
            "bold": False,
            "horizontal": "left",
            "vertical": "center",
            "apply_border": True,
            "column_width": {"C": 16},
        },
    }
    a.apply_action(action)
    first = (copy(a.workbook["Sales"]["C2"].font), copy(a.workbook["Sales"]["C2"].alignment), a.workbook["Sales"].column_dimensions["C"].width)
    a.apply_action(action)
    second = (copy(a.workbook["Sales"]["C2"].font), copy(a.workbook["Sales"]["C2"].alignment), a.workbook["Sales"].column_dimensions["C"].width)
    assert first[0] == second[0]
    assert first[1] == second[1]
    assert first[2] == second[2]


def test_xlsx_case_03_unify_style_on_inventory(tmp_path: Path):
    p = _create_sample_inventory(tmp_path / "inventory.xlsx")
    a = XlsxAdapter(str(p))
    a.apply_action({"action_type": "unify_style", "target": {"sheet": "Inventory"}, "params": {}})

    ws = a.workbook["Inventory"]
    assert ws["A1"].font.bold is True
    assert ws["A1"].alignment.horizontal == "center"
    assert ws["A2"].font.bold is False
    assert ws["A2"].border.left.style == "thin"


def test_xlsx_case_04_unify_style_idempotent(tmp_path: Path):
    p = _create_sample_inventory(tmp_path / "inventory.xlsx")
    a = XlsxAdapter(str(p))
    action = {"action_type": "unify_style", "target": {"sheet": "Inventory"}, "params": {}}
    a.apply_action(action)
    first_border = copy(a.workbook["Inventory"]["B3"].border)
    first_font = copy(a.workbook["Inventory"]["B3"].font)
    a.apply_action(action)
    second_border = copy(a.workbook["Inventory"]["B3"].border)
    second_font = copy(a.workbook["Inventory"]["B3"].font)
    assert first_border == second_border
    assert first_font == second_font


def test_xlsx_case_05_extract_content_with_columns_and_where(tmp_path: Path):
    p = _create_sample_sales(tmp_path / "sales.xlsx")
    a = XlsxAdapter(str(p))
    result = a.apply_action(
        {
            "action_type": "extract_content",
            "target": {"sheet": "Sales"},
            "params": {"columns": ["姓名", "销售额"], "where": {"column": "状态", "op": "==", "value": "在售"}},
        }
    )
    assert result.success is True
    assert result.details["row_count"] == 2
    assert result.details["rows"][0]["姓名"] == "张三"
    assert "销售额" in result.details["columns"]


def test_xlsx_case_06_extract_content_condition_string_with_stats(tmp_path: Path):
    p = _create_sample_sales(tmp_path / "sales.xlsx")
    a = XlsxAdapter(str(p))
    result = a.apply_action(
        {
            "action_type": "extract_content",
            "target": {"sheet": "Sales"},
            "params": {"columns": ["销售额", "地区"], "condition": "销售额 > 100"},
        }
    )
    assert result.success is True
    assert result.details["row_count"] == 2
    assert "均值" in result.details["stats"]


def test_xlsx_case_07_reorder_sort_rows_desc(tmp_path: Path):
    p = _create_sample_sales(tmp_path / "sales.xlsx")
    a = XlsxAdapter(str(p))
    result = a.apply_action(
        {
            "action_type": "reorder_paragraphs",
            "target": {"sheet": "Sales"},
            "params": {"sort_by": "销售额", "order": "desc", "header_row": 1},
        }
    )
    ws = a.workbook["Sales"]
    assert result.success is True
    assert ws["A2"].value == "王五"
    assert ws["C2"].value == 150


def test_xlsx_case_08_reorder_move_region(tmp_path: Path):
    p = _create_sample_region(tmp_path / "region.xlsx")
    a = XlsxAdapter(str(p))
    result = a.apply_action(
        {
            "action_type": "reorder_paragraphs",
            "target": {"sheet": "Region"},
            "params": {"move_range": "3:4", "to_row": 2},
        }
    )
    ws = a.workbook["Region"]
    assert result.success is True
    assert ws["A2"].value == "华北"
    assert ws["A3"].value == "华南"


def test_xlsx_case_09_auto_column_width_and_idempotent(tmp_path: Path):
    p = _create_sample_inventory(tmp_path / "inventory.xlsx")
    a = XlsxAdapter(str(p))
    action = {"action_type": "auto_column_width", "target": {"sheet": "Inventory"}, "params": {"padding": 2}}
    a.apply_action(action)
    w1 = a.workbook["Inventory"].column_dimensions["A"].width
    a.apply_action(action)
    w2 = a.workbook["Inventory"].column_dimensions["A"].width
    assert w1 == w2
    assert w1 >= 8


def test_xlsx_case_10_freeze_header_row(tmp_path: Path):
    p = _create_sample_sales(tmp_path / "sales.xlsx")
    a = XlsxAdapter(str(p))
    result = a.apply_action({"action_type": "freeze_header_row", "target": {"sheet": "Sales"}, "params": {"header_row": 1}})
    assert result.success is True
    assert a.workbook["Sales"].freeze_panes == "A2"
